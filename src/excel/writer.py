"""
Excel writer for SO CONG VAN DEN-LIENDO.xlsx.

Sheet layout:
  DATA  — business rows (one per processed email)
  META  — dedup/run metadata (one per processed email, mirrors _processed.json)

Safe-write strategy:
  Load → append row → save.
  If the file is locked (open in Excel), ExcelLockedError is raised so the
  caller can prompt the user to close it and retry.

Column order is defined by DATA_COLUMNS / META_COLUMNS lists.
To add a new column, append to the appropriate list — existing files will
receive the new column only on the next write (openpyxl preserves old rows).
"""
from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


class ExcelLockedError(PermissionError):
    """Raised when the target Excel file is locked (open in another process)."""

    def __init__(self, path: Path) -> None:
        super().__init__(f"File Excel đang được mở bởi tiến trình khác: {path}")
        self.excel_path = path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Column definitions ─────────────────────────────────────────────────────

# Simple keys used as dict keys in row_data passed to append_data_row.
DATA_COLUMNS: List[str] = [
    "Ngày nhận công văn",
    "Số công văn",
    "Loại công văn",
    "Ngày issue công văn",
    "Số tháng deadline",
    "Deadline trả lời Cục",
    "Số đơn",
    "Loại hình đơn",
    "Nội dung công văn",
    "Người thực hiện",
    "Ký nhận",
    "Ghi chú",
    "Số biên lai",
    "Sô bằng",
    "Lỗi",
]

# Display headers written to the worksheet header row (may contain \n sub-notes).
DATA_COLUMN_HEADERS: List[str] = [
    "Ngày nhận công văn",
    "Số công văn",
    "Loại công văn",
    "Ngày issue công văn\n(MM/DD/YYYY)",
    "Số tháng deadline\n (kéo công thức xuống)",
    "Deadline trả lời Cục (nếu có)\n(MM/DD/YYYY)\n(kéo công thức xuống)",
    "Số đơn",
    "Loại hình đơn",
    "Nội dung công văn",
    "Người thực hiện",
    "Ký nhận",
    "Ghi chú",
    "Số biên lai",
    "Sô bằng",
    "Lỗi",
]

META_COLUMNS: List[str] = [
    "message_id",
    "internet_message_id",
    "date_folder",
    "so_don",
    "attachment_filenames",
    "processed_at",
    "run_status",
]

# ── Styles ─────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
_ROW_ERROR_FILL  = PatternFill("solid", fgColor="FFCCCC")   # light red — data validation errors
_ROW_SCAN_FILL   = PatternFill("solid", fgColor="FFFF00")   # yellow — scanned PDF, needs review

# Columns that are never auto-populated from PDF parsing (user fills manually).
# These are hidden in the exported sheet to reduce clutter but can be unhidden in Excel.
_HIDDEN_COLUMNS: set[str] = set()


# ── ExcelWriter ────────────────────────────────────────────────────────────

class ExcelWriter:
    """Creates and updates the daily Excel file for công văn records."""

    def __init__(self, daily_folder: Path, excel_filename: str = "SO CONG VAN DEN-LIENDO.xlsx"):
        self.excel_path = daily_folder / excel_filename

    def reset(self) -> None:
        """Remove the previous workbook so the current scan rebuilds it from scratch."""
        lock_file = self.excel_path.parent / f"~${self.excel_path.name}"
        if lock_file.exists():
            raise ExcelLockedError(self.excel_path)
        if not self.excel_path.exists():
            return
        try:
            self.excel_path.unlink()
        except PermissionError as exc:
            raise ExcelLockedError(self.excel_path) from exc

    def append_data_row(self, row_data: Dict[str, Any],
                        highlight_red: bool = False,
                        highlight_yellow: bool = False) -> None:
        """Append one row to the DATA sheet.

        Uses DATA_COLUMNS key order to map row_data values to columns,
        so column positions are always stable regardless of header display text.
        Unknown keys in row_data are silently skipped.

        Args:
            row_data:         Dict mapping DATA_COLUMNS keys to values.
            highlight_red:    If True, apply a light-red fill (missing required fields).
            highlight_yellow: If True, apply a yellow fill (scanned PDF — needs review).
                              Takes lower priority than highlight_red.
        """
        wb = _load_or_create(self.excel_path)
        ws = wb["DATA"]
        row_idx = _next_empty_row(ws, len(DATA_COLUMNS))
        for col_idx, col_key in enumerate(DATA_COLUMNS, start=1):
            value = row_data.get(col_key)
            if value is not None:
                cell = ws.cell(row=row_idx, column=col_idx, value=_coerce(value))
                cell.alignment = _CELL_ALIGN
                # Date objects need an explicit format so Excel renders them as dates (not serials)
                if isinstance(value, (date, datetime)):
                    cell.number_format = "MM/DD/YYYY"

        # Always write EDATE formula to col F — relative refs adjust automatically on paste to master
        _deadline_col = DATA_COLUMNS.index("Deadline trả lời Cục") + 1
        fcell = ws.cell(row=row_idx, column=_deadline_col,
                        value=f'=IFERROR(EDATE(D{row_idx},E{row_idx}),"0DL")')
        fcell.number_format = "MM/DD/YYYY"
        fcell.alignment = _CELL_ALIGN

        if highlight_red:
            for col_idx in range(1, len(DATA_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ROW_ERROR_FILL
        elif highlight_yellow:
            for col_idx in range(1, len(DATA_COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = _ROW_SCAN_FILL
        self._save(wb)
        logger.debug("DATA row %d written to %s", row_idx, self.excel_path.name)

    def append_date_row(self, date_str: str) -> None:
        """Write a date separator row: col A = date, all other columns blank.

        This row groups documents by date (as seen in the reference Excel).
        """
        wb = _load_or_create(self.excel_path)
        ws = wb["DATA"]
        row_idx = _next_empty_row(ws, len(DATA_COLUMNS))
        cell = ws.cell(row=row_idx, column=1, value=date_str)
        cell.alignment = _CELL_ALIGN
        self._save(wb)
        logger.debug("Date row written: %s at row %d", date_str, row_idx)

    def next_sequence_number(self) -> int:
        """Return the next sequence number (1-based) for data rows.

        Counts only rows where col A holds an integer value (data rows),
        ignoring date separator rows (where col A is a date string).
        Returns 1 if the file does not exist or has no data rows yet.
        """
        if not self.excel_path.exists():
            return 1
        try:
            wb = _load_or_create(self.excel_path)
            ws = wb["DATA"]
            max_seq = 0
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=1).value
                if isinstance(val, (int, float)) and float(val) == int(float(val)):
                    max_seq = max(max_seq, int(float(val)))
            return max_seq + 1
        except Exception as exc:
            logger.warning("Could not read sequence number from %s: %s", self.excel_path.name, exc)
            return 1

    def append_meta_row(self, meta_data: Dict[str, Any]) -> None:
        """Append one row to the META sheet. Keys must match META_COLUMNS."""
        wb = _load_or_create(self.excel_path)
        ws = wb["META"]
        row_idx = _next_empty_row(ws, len(META_COLUMNS))
        for col_idx, col_name in enumerate(META_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_coerce(meta_data.get(col_name)))
        self._save(wb)

    def _save(self, wb: Workbook) -> None:
        """
        Save the workbook.

        Before writing, proactively checks for the  ~$<filename>  lock file
        that Excel creates whenever a workbook is open — this is faster and
        more reliable than waiting for openpyxl to raise PermissionError.
        Raises ExcelLockedError if the file is locked by another process.
        """
        # Proactive lock-file check (Excel on Windows always creates ~$<name>)
        lock_file = self.excel_path.parent / f"~${self.excel_path.name}"
        if lock_file.exists():
            raise ExcelLockedError(self.excel_path)

        try:
            wb.save(self.excel_path)
        except PermissionError as exc:
            raise ExcelLockedError(self.excel_path) from exc


# ── Helpers ────────────────────────────────────────────────────────────────

def _load_or_create(excel_path: Path) -> Workbook:
    """Load existing workbook or create a new one with DATA + META sheets."""
    if excel_path.exists():
        try:
            wb = openpyxl.load_workbook(excel_path)
        except Exception as exc:
            raise IOError(
                f"Cannot open Excel file {excel_path}: {exc}\n"
                "If the file is corrupted, rename or delete it to start fresh."
            ) from exc
        if "DATA" not in wb.sheetnames:
            _add_sheet(wb, "DATA", DATA_COLUMNS, position=0, display_headers=DATA_COLUMN_HEADERS)
        if "META" not in wb.sheetnames:
            _add_sheet(wb, "META", META_COLUMNS)
        return wb

    wb = Workbook()
    # Remove default sheet
    for name in wb.sheetnames:
        del wb[name]
    _add_sheet(wb, "DATA", DATA_COLUMNS, display_headers=DATA_COLUMN_HEADERS)
    _add_sheet(wb, "META", META_COLUMNS)
    return wb


_COLUMN_WIDTHS: Dict[str, int] = {
    "Ngày nhận công văn": 20,
    "Số công văn": 14,
    "Loại công văn": 16,
    "Ngày issue công văn": 22,
    "Số tháng deadline": 22,
    "Deadline trả lời Cục": 30,
    "Số đơn": 22,
    "Loại hình đơn": 16,
    "Nội dung công văn": 50,
    "Người thực hiện": 20,
    "Ký nhận": 12,
    "Ghi chú": 30,
    "Số biên lai": 16,
    "Sô bằng": 16,
    "Lỗi": 40,
}


def _add_sheet(wb: Workbook, name: str, columns: List[str], position: int = None,
               display_headers: List[str] = None) -> None:
    ws = wb.create_sheet(name, position)
    ws.row_dimensions[1].height = 45
    headers = display_headers if display_headers is not None else columns
    for col_idx, (col_key, col_header) in enumerate(zip(columns, headers), start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        col_letter = get_column_letter(col_idx)
        col_dim = ws.column_dimensions[col_letter]
        col_dim.width = _COLUMN_WIDTHS.get(col_key, max(len(col_key) + 4, 18))
        if col_key in _HIDDEN_COLUMNS:
            col_dim.hidden = True
    ws.freeze_panes = "A2"


def _next_empty_row(ws, num_cols: int) -> int:
    """Return the 1-indexed row number of the next empty row after headers."""
    if ws.max_row <= 1:
        return 2
    # Walk backward to find the last non-empty row
    for row in range(ws.max_row, 1, -1):
        if any(ws.cell(row=row, column=c).value is not None for c in range(1, num_cols + 1)):
            return row + 1
    return 2


def _coerce(value: Any) -> Any:
    """Pass values through; date/datetime objects are written natively so Excel formulas work."""
    return value


def format_date(d: Optional[date]) -> str:
    """Format a date as DD/MM/YYYY string, or empty string if None."""
    if d is None:
        return ""
    return d.strftime("%d/%m/%Y")

