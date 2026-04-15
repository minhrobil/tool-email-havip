"""Dump full Excel structure including styles, merges, column widths."""
import openpyxl
from openpyxl.utils import get_column_letter

path = r"C:\Users\minh.nguyenq3\Documents\QuangMinh\Work\Code\HomeX\Repositories\mail-extract\docs\SO CONG VAN DEN-LIENDO.xlsx"
wb = openpyxl.load_workbook(path)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ===")

    # Merged cells
    if ws.merged_cells.ranges:
        print(f"  Merged cells: {[str(m) for m in ws.merged_cells.ranges]}")

    # Freeze panes
    print(f"  Freeze panes: {ws.freeze_panes}")

    # Row heights
    for row_idx, rd in ws.row_dimensions.items():
        if rd.height:
            print(f"  Row {row_idx} height: {rd.height}")

    # Column widths + header styles
    print(f"\n  --- Columns ---")
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        width = ws.column_dimensions[letter].width if letter in ws.column_dimensions else None
        header_cell = ws.cell(row=1, column=c)
        font = header_cell.font
        fill = header_cell.fill
        align = header_cell.alignment
        print(f"  Col {letter} (idx={c}): width={width}")
        print(f"    Header value : {repr(header_cell.value)}")
        print(f"    Font         : bold={font.bold}, color={font.color.rgb if font.color and font.color.type=='rgb' else '?'}, size={font.size}, name={font.name}")
        print(f"    Fill         : type={fill.fill_type}, fgColor={fill.fgColor.rgb if fill.fgColor and fill.fgColor.type=='rgb' else '?'}")
        print(f"    Alignment    : h={align.horizontal}, v={align.vertical}, wrap={align.wrap_text}")

    # Data rows (first 3)
    print(f"\n  --- Data rows (first 3) ---")
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=4), start=2):
        for cell in row:
            if cell.value is not None:
                align = cell.alignment
                print(f"  [{r_idx},{cell.column}] value={repr(cell.value)}, wrap={align.wrap_text}, v={align.vertical}")

